import time
import cv2
import numpy as np
import requests
import os
from openpyxl import Workbook

BOOK_NAME = 'book6'

wb = Workbook()
ws = wb.active

all_tasks = []

cctr = 1
for ctr in range(1, len(os.listdir("C:/Users/vince/Desktop/{}_t/".format(BOOK_NAME)))+1):
    f_name = 'C:\\Users\\vince\\Desktop\\{}_t\\{}.txt'.format(BOOK_NAME, ctr)
    print(f_name)
    f = open(f_name, "r")
    m_str = f.read()

    m_tasks = m_str.split("\n\n\n\n")
    m_tasks = m_tasks[0:len(m_tasks)-1]

    for tsk in m_tasks:
        tsk = tsk.replace("\n", "")
        if tsk is not None and tsk is not "":
            c_a = 'A{}'.format(cctr)
            c_b = 'B{}'.format(cctr)
            ws[c_a] = str(cctr)
            ws[c_b] = str(tsk)
            cctr += 1
    f.close()

wb.save("C:/Users/vince/Desktop/output{}.xlsx".format(BOOK_NAME))
