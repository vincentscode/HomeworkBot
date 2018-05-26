import os
import numpy as np
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

task_ctr = []

all_tasks = []

full_str = ""
for ctr in range(1, len(os.listdir("C:/Users/vince/Desktop/book5_t/"))+1):
    f_name = 'C:\\Users\\vince\\Desktop\\book5_t\\{}.txt'.format(ctr)
    print(f_name)
    f = open(f_name, "r")
    m_str = f.read()

    m_tasks = m_str.split("\n\n\n\n")
    m_tasks = m_tasks[0:-1]
    task_ctr.append(len(m_tasks))

    for tsk in m_tasks:
        all_tasks.append(m_tasks)

    full_str += m_str
    f.close()

n_tasks = []
for tsk in all_tasks:
    for t in tsk:
        t = t.replace("\n", "")
        n_tasks.append(t)

for i in range(1, len(n_tasks)):
    c_a = 'A{}'.format(i)
    c_b = 'B{}'.format(i)
    ws[c_a] = str(i)
    ws[c_b] = n_tasks[i]

wb.save("C:/Users/vince/Desktop/output5.xlsx")

print("Avg tasks per page:", np.mean(task_ctr))
print("Task_count:", len(n_tasks))