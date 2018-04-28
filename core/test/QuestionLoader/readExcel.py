import openpyxl
wb = openpyxl.load_workbook('C:\\Users\\vince\\Desktop\\math_questions.xlsx')
ws = wb[wb.sheetnames[0]]
print(ws['B1'].value)
print(tuple(ws.columns)[0])