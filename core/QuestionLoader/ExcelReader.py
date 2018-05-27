import openpyxl
wb = openpyxl.load_workbook('C:\\Users\\vince\\Desktop\\output5.xlsx')
ws = wb[wb.sheetnames[0]]

tasks = [ws['B{}'.format(i)].value for i in range(1, 4000) if ws['B{}'.format(i)].value is not None]
print(tasks)
print(len(tasks))