import time
import cv2
import numpy as np
import requests
import os
from openpyxl import Workbook

BOOK_NAME = 'book7'

# keys = 4beb31282988957, acf7c56e6088957

requests_ctr = 0

os.mkdir("C:/Users/vince/Desktop/{}_t/".format(BOOK_NAME))

for ctr in range(1, len(os.listdir("C:/Users/vince/Desktop/{}_c/".format(BOOK_NAME))) + 1):
    # Settings
    method = cv2.TM_CCOEFF_NORMED
    f_name = 'C:/Users/vince/Desktop/{]_c/{}.jpg'.format(BOOK_NAME, ctr)
    print(f_name)
    shorten_output = True

    # Image
    img_rgb = cv2.imread(f_name)
    offset_left = 100
    offset_top = 265
    img_rgb = img_rgb[offset_left:1920, offset_top:1080]

    img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)

    # noinspection PyShadowingNames
    class Rect:
        def __init__(self, x, y, w, h):
            self.x = x
            self.y = y
            self.w = w
            self.h = h

        def __str__(self):
            if not shorten_output:
                return "<Rect: x = " + str(self.x) + "; y = " + str(self.y) + "; w = " + str(self.w) + "; h = " + str(
                    self.h) + ">"
            else:
                return "<Rect: x = " + str(self.x) + "; y = " + str(self.y) + ">"

        def __repr__(self):
            return self.__str__()


    # OCR
    def send_ocr(filename, overlay=False, api_key='4beb31282988957', language='ger', pdf=False):
        global requests_ctr
        try:
            print("Requests: %r" % requests_ctr)
            requests_ctr += 1
            payload = {'isOverlayRequired': overlay,
                       'apikey': api_key,
                       'language': language,
                       'isCreateSearchablePdf': pdf
                       }
            # noinspection PyShadowingNames
            with open(filename, 'rb') as f:
                r = requests.post('https://api.ocr.space/parse/image',
                                  files={filename: f},
                                  data=payload,
                                  )
            return r.json()
        except requests.exceptions.ConnectionError:
            print("Connection blocked... Waiting")
            time.sleep(10)
            return send_ocr(filename, overlay, api_key, language, pdf)


    # Template matching
    # noinspection PyShadowingNames
    def find(template, threshold=0.85):
        w, h = template.shape[::-1]
        res = cv2.matchTemplate(img_gray, template, method)
        loc = np.where(res >= threshold)
        rects = []
        for pt in zip(*loc[::-1]):
            rects.append(Rect(pt[0], pt[1], w, h))
        return rects


    # Numbers
    find_left = find(cv2.imread("/patterns/number.png", 0), 0.7)

    # Parts
    lines_left = []
    for bx in find_left:
        if not lines_left:
            lines_left.append([bx])
        else:
            if lines_left[-1][-1].y - 10 < bx.y < lines_left[-1][-1].y + 10:
                lines_left[-1].append(bx)
            else:
                lines_left.append([bx])

    lines_cleaned = [None] * len(lines_left)
    for i in range(len(lines_cleaned)):
        # noinspection PyRedeclaration
        last_box = None
        for bx in lines_left[i]:
            if not last_box:
                # noinspection PyTypeChecker
                lines_cleaned[i] = [bx]
            else:
                if not last_box.x - 10 < bx.x < last_box.x + 10:
                    lines_cleaned[i].append(bx)
            last_box = bx

    numbers = []
    for itm in lines_cleaned:
        # noinspection PyTypeChecker
        for x in itm:
            numbers.append(x)
            cv2.circle(img_rgb, (x.x, x.y + 40), 3, (0, 255, 0), -1)

    print(str(numbers).replace(", ", ",\n"))

    tasks = []
    for i in range(len(numbers)):
        cur = numbers[i]
        if i != len(numbers) - 1:
            nxt = numbers[i + 1]
        else:
            nxt = Rect(0, 1406, 0, 0)
        start = cur.x, cur.y + 40
        end = 800, nxt.y - 5
        cv2.rectangle(img_rgb, start, end, (255, 0, 0))
        tasks.append(img_rgb[start[1]:end[1], start[0]:end[0]])

    for t in range(len(tasks)):
        ff = '/temp/Task {}.jpg'.format(t)
        cv2.imwrite(ff, tasks[t])

    print("Wrote")
    final_string = ""
    # noinspection PyBroadException
    try:
        for t in range(len(tasks)):
            ff = '/temp/Task {}.jpg'.format(t)
            print(ff)
            res = send_ocr(filename=ff, language='ger', overlay=False, pdf=False)
            if res and res["ParsedResults"] and res["ParsedResults"][0]:
                final_string += res["ParsedResults"][0]["ParsedText"] + "\n\n\n\n"
        file = open("C:/Users/vince/Desktop/{}_t/{}.txt".format(BOOK_NAME, ctr), "w")
        file.write(final_string)
        file.close()
    except Exception:
        # noinspection PyUnboundLocalVariable
        print(res)
        exit(-1)

# CREATE EXCEL
wb = Workbook()
ws = wb.active

task_ctr = []

all_tasks = []

full_str = ""
for ctr in range(1, len(os.listdir("C:/Users/vince/Desktop/{}_t/".format(BOOK_NAME)))+1):
    f_name = 'C:\\Users\\vince\\Desktop\\{}_t\\{}.txt'.format(BOOK_NAME, ctr)
    print(f_name)
    f = open(f_name, "r")
    m_str = f.read()

    m_tasks = m_str.split("\n\n\n\n")
    m_tasks = m_tasks[0:-1]
    task_ctr.append(len(m_tasks))

    for tsk in m_tasks:
        all_tasks.append(m_tasks)

    full_str += m_str
    f.close()

n_tasks = []
for tsk in all_tasks:
    for t in tsk:
        t = t.replace("\n", "")
        n_tasks.append(t)

for i in range(1, len(n_tasks)):
    c_a = 'A{}'.format(i)
    c_b = 'B{}'.format(i)
    ws[c_a] = str(i)
    ws[c_b] = n_tasks[i]

wb.save("C:/Users/vince/Desktop/output{}.xlsx".format(BOOK_NAME))

print("Avg tasks per page:", np.mean(task_ctr))
print("Task_count:", len(n_tasks))
